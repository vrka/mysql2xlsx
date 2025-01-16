#!/bin/python3

from copy import copy

import click
import mysql.connector
from openpyxl import Workbook, load_workbook, utils, workbook
from openpyxl import __version__ as openpyxl_version
from packaging import version

VERSION = "1.4"


@click.command()
@click.option('-u', '--user', help="username")
@click.option('-p', '--password', help="password")
@click.option('-h', '--hostname', help="MySql server hostname")
@click.option('-P', '--port', default=3306, help="MySql server port")
@click.option('-d', '--database', help="database name")
@click.option('-o', '--output', type=click.Path(writable=True), default='output.xlsx', help="xlsx filename")
@click.option('-t', '--template', type=click.Path(exists=True, readable=True), help="xlsx template filename")
@click.option('-f', '--file', type=click.Path(exists=True, readable=True), help="SQL file path")
@click.argument('sql', required=False)
@click.version_option(VERSION)
@click.version_option(VERSION, '--version-simple', message="%(version)s",
                      help="Show the version number (only) and exit")
def main(user, password, hostname, port, database, output, template, file, sql):
    """Saves output of SQL command as XLSX file, optionally formatted as template file"""
    if file:
        with open(file, 'r') as f:
            sql = f.read()

    if not sql:
        raise click.UsageError("You must provide an SQL command or specify a file using -f/--file.")

    db = mysql.connector.connect(user=user, password=password, host=hostname, port=port, database=database)
    cur = db.cursor()
    cur.execute(sql)

    if template:
        wb = load_workbook(template)
        ws = wb.worksheets[0]
        name2index = {val: idx for idx, val in enumerate(cur.column_names)}
        name2col = {}

        row_src = ws[ws.freeze_panes].row if ws.freeze_panes else 2
        row_idx = row_src + 1

        for row in cur.fetchall():
            ws.insert_rows(row_idx)
            for col_idx in range(1, ws.max_column + 1):
                # copy format from row row_src
                src = ws.cell(row_src, col_idx)
                dst = ws.cell(row_idx, col_idx)
                dst._style = copy(src._style)
                if src.value[0] == "_":
                    dst.value = row[name2index[src.value[1:]]]
                    name2col[src.value[1:]] = col_idx
            row_idx += 1
        rules = list(ws.conditional_formatting._cf_rules)
        for cf_rule in rules:
            rng = copy(list(cf_rule.cells.ranges)[0])
            rng.max_row = row_idx - 1
            for rule in ws.conditional_formatting[cf_rule]:
                ws.conditional_formatting.add(str(rng), rule)
        # process last row (replace range in formulas)
        for name in name2col:
            col_name = utils.cell.get_column_letter(name2col[name])
            named_range = f"{ws.title}!${col_name}${row_src}:${col_name}${row_idx - 2}"
            new_range = workbook.defined_name.DefinedName('data_' + name, attr_text=named_range)
            if version.parse(openpyxl_version) < version.parse("3.1.0"):
                wb.defined_names.append(new_range)
            else:
                wb.defined_names[f"data_{name}"] = new_range

        ws.delete_rows(row_src)
    else:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.append(cur.column_names)

        for row in cur.fetchall():
            ws.append(row)

    wb.save(output)


if __name__ == '__main__':
    main()
